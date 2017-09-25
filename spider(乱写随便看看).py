# -*- coding:utf-8 -*-
import requests
from bs4 import BeautifulSoup
import re,time,sys
from openpyxl import Workbook
def getHTMLText(url,k):
    try:
        if (k == 0):
            kw = {}
        else:
            kw = {'start': k, 'filter': ''}
        r = requests.get(url, params=kw, headers={'User-Agent': 'Mozilla/4.0'})
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text
    except:
        print("Failed!")
def getData():
    basicUrl = 'https://movie.douban.com/top250'
    k = 0
    data = []
    while k <= 250:
        html = getHTMLText(basicUrl, k)
        time.sleep(2)
        soup = BeautifulSoup(html, "html.parser")
        movieList=soup.find('ol',attrs={'class':'grid_view'})#找到第一个class属性值为grid_view的ol标签
        for movieLi in movieList.find_all('li'):#找到所有li标签
            #得到电影名字
            movieHd=movieLi.find('div',attrs={'class':'hd'})#找到第一个class属性值为hd的div标签
            movieName=movieHd.find('span',attrs={'class':'title'}).getText()#找到第一个class属性值为title的span标签
                                                                               #也可使用.string方法
            #得到电影的评分
            movieScore=movieLi.find('span',attrs={'class':'rating_num'}).getText()
            #得到电影的评价人数
            movieEval=movieLi.find('div',attrs={'class':'star'})
            movieEvalNum=re.findall(r'\d+',str(movieEval))[-1]
            # 得到电影的短评
            movieQuote = movieLi.find('span', attrs={'class': 'inq'})
            if(movieQuote):
                data.append([movieName,movieScore,movieEvalNum,movieQuote.getText()])
            else:
                data.append([movieName, movieScore, movieEvalNum, '无'])
        k += 25
        if(k>=250):
            return data

def print_movie_lists_excel(data):
    wb = Workbook()
    ws = []
    ws.append(wb.create_sheet(title='data'))
    ws[0].append(['序号', '影片名','评分', '评论人数', '短评'])
    count = 1
    for bl in data:
        ws[0].append([count, bl[0], bl[1],
                          bl[2],bl[3]])
        count += 1
    save_path = 'movie.xlsx'
    wb.save(save_path)
    print 'Downloading finished'

data= getData()
print_movie_lists_excel(data)
